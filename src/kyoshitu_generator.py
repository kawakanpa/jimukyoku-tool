"""
教室案内表 Excel生成モジュール
- 1日1シート、指定期間分を1ファイルに出力
- シンプルな表形式
"""
import os
from datetime import date, timedelta
from itertools import groupby

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schedule_builder import ScheduleMap, get_entry_label

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

HEADER_FILL = PatternFill("solid", fgColor="FFB6C1")   # ピンク
ROOM_FILL   = PatternFill("solid", fgColor="FFDDE1")   # 薄ピンク
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def generate_kyoshitu(
    schedule: ScheduleMap,
    settings: dict,
    start_date: date,
    end_date: date,
    output_dir: str,
) -> str:
    """
    指定期間の教室案内表Excelを生成してパスを返す。
    """
    floor_rooms: dict = settings.get("floor_rooms", {"6階": [601, 602, 603]})
    slots_order: list[str] = settings.get("time_slots_order", [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    start = date(int(start_date.year), int(start_date.month), int(start_date.day))
    end   = date(int(end_date.year),   int(end_date.month),   int(end_date.day))
    days_count = (end - start).days + 1
    for i in range(days_count):
        current = start + timedelta(days=i)
        wd = WEEKDAY_JP[current.weekday()]
        ws = wb.create_sheet(title=f"{current.month}/{current.day}({wd})")
        _write_sheet(ws, current, wd, floor_rooms, slots_order, schedule,
                     current.year, current.month, current.day)

    label = f"{start_date.strftime('%Y.%m.%d')}-{end_date.strftime('%m.%d')}"
    filename = f"教室案内表_{label}.xlsx"
    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    return out_path


def _write_sheet(ws, d, wd, floor_rooms, slots_order, schedule, year, month, day):
    day_schedule = schedule.get(d, {})

    # --- 使用する時間帯のみに絞る ---
    used_slots = set()
    for room_slots in day_schedule.values():
        used_slots.update(room_slots.keys())
    display_slots = [s for s in slots_order if s in used_slots] or slots_order

    # === タイトル行 ===
    title = f"教室案内表　{year}年{month}月{day}日（{wd}）"
    ws.cell(1, 1, title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + len(display_slots))

    # === ヘッダー行 ===
    headers = ["階", "教室"] + display_slots
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(2, col, h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # === データ行 ===
    row = 3
    for floor_name, rooms in floor_rooms.items():
        floor_start = row
        for room in rooms:
            # 階列（最初の部屋だけ後でマージ）
            ws.cell(row, 1, floor_name if room == rooms[0] else None)
            ws.cell(row, 1).fill = ROOM_FILL
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, 1).border = BORDER

            # 教室番号
            c = ws.cell(row, 2, room)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = ROOM_FILL
            c.border = BORDER

            # 各時間帯
            for col, slot in enumerate(display_slots, start=3):
                entries = day_schedule.get(room, {}).get(slot, [])
                label = get_entry_label(entries) if entries else ""
                cell = ws.cell(row, col, label)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border = BORDER
            row += 1

        # 階列をマージ
        if len(rooms) > 1:
            ws.merge_cells(start_row=floor_start, start_column=1,
                           end_row=row - 1, end_column=1)

    # === フッター ===
    note = "※自習室利用時間　★平日11:00～20:30（平日夜間講義がない日～18:30）　★土日祝11:00～17:00"
    ws.cell(row, 1, note)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2 + len(display_slots))
    ws.cell(row, 1).font = Font(size=9)

    # === 列幅 ===
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 7
    for i in range(len(display_slots)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 20

    # === 行高 ===
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    for r in range(3, row):
        ws.row_dimensions[r].height = 45
