"""
鳥瞰図 Excel生成モジュール
- 週ごとのブロック（月〜日）
- 行：時間帯、列：曜日×教室
- シンプル版（1シート/年or期間）
"""
import os
from collections import defaultdict
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schedule_builder import ScheduleMap, get_entry_label

WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

# スタイル定義
THIN  = Side(style="thin")
THICK = Side(style="medium")
BORDER_ALL  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_WEEK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

FILL_WEEKHDR = PatternFill("solid", fgColor="C6EFCE")   # 週ヘッダー（緑系）
FILL_SLOTHDR = PatternFill("solid", fgColor="DDEBF7")   # 時間帯ヘッダー（青系）
FILL_ROOMHDR = PatternFill("solid", fgColor="FFF2CC")   # 教室ヘッダー（黄系）
FILL_SAT     = PatternFill("solid", fgColor="E2EFDA")   # 土曜列（薄緑）
FILL_SUN     = PatternFill("solid", fgColor="FCE4D6")   # 日曜列（薄橙）


def generate_kanzu(
    schedule: ScheduleMap,
    settings: dict,
    start_date: date,
    end_date: date,
    output_dir: str,
    label: str = "",
) -> str:
    """
    鳥瞰図Excelを生成してパスを返す。
    label: ファイル名用ラベル (例: "2026.5-8月")
    """
    floor_rooms: dict = settings.get("floor_rooms", {"6階": [601, 602, 603]})
    all_rooms: list[int] = []
    for rooms in floor_rooms.values():
        all_rooms.extend(rooms)

    slots_order: list[str] = settings.get("time_slots_order", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "鳥瞰図"
    ws.freeze_panes = "B1"

    weeks = _split_weeks(start_date, end_date)

    # シート全体を構築
    current_row = 1
    for week_start, week_days in weeks:
        current_row = _write_week_block(
            ws, week_days, all_rooms, slots_order, schedule, current_row
        )
        current_row += 1  # 週間の空白行

    # 列幅設定
    ws.column_dimensions["A"].width = 8   # 時間帯列
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    if not label:
        label = f"{start_date.strftime('%Y.%m')}-{end_date.strftime('%m月')}"
    filename = f"鳥瞰図_{label}.xlsx"
    out_path = os.path.join(output_dir, filename)
    wb.save(out_path)
    return out_path


def _split_weeks(start: date, end: date) -> list[tuple[date, list[date]]]:
    """開始日〜終了日を週単位（月曜起点）に分割する"""
    weeks = []
    # 月曜日に戻す
    cur = start - timedelta(days=start.weekday())
    while cur <= end:
        week_days = [cur + timedelta(days=i) for i in range(7)]
        weeks.append((cur, week_days))
        cur += timedelta(weeks=1)
    return weeks


def _write_week_block(
    ws, week_days: list[date], rooms: list[int],
    slots_order: list[str], schedule: ScheduleMap, start_row: int
) -> int:
    """1週分のブロックを書き込み、次の開始行を返す"""
    row = start_row

    # ─── 行1: 曜日ヘッダー ───
    ws.cell(row, 1, "").border = BORDER_ALL
    col = 2
    for d in week_days:
        wd = WEEKDAY_JP[d.weekday()]
        date_str = f"{d.month}/{d.day}（{wd}）"
        # 1日につきrooms列分マージ
        if len(rooms) > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + len(rooms) - 1
            )
        cell = ws.cell(row, col, date_str)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        fill = FILL_SAT if d.weekday() == 5 else (FILL_SUN if d.weekday() == 6 else FILL_WEEKHDR)
        cell.fill = fill
        cell.border = BORDER_ALL
        col += len(rooms)
    row += 1

    # ─── 行2: 教室番号ヘッダー ───
    ws.cell(row, 1, "時間帯").border = BORDER_ALL
    ws.cell(row, 1).fill = FILL_SLOTHDR
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 1).font = Font(bold=True)
    col = 2
    for d in week_days:
        for room in rooms:
            cell = ws.cell(row, col, str(room))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = FILL_ROOMHDR
            cell.font = Font(bold=True)
            cell.border = BORDER_ALL
            col += 1
    row += 1

    # ─── 時間帯行 ───
    for slot in slots_order:
        # 当週にこのスロットが使われているか確認
        has_data = any(
            d in schedule and room in schedule[d] and slot in schedule[d][room]
            for d in week_days for room in rooms
        )
        if not has_data:
            continue

        cell = ws.cell(row, 1, slot)
        cell.fill = FILL_SLOTHDR
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_ALL

        col = 2
        for d in week_days:
            wd_idx = d.weekday()
            for room in rooms:
                entries = schedule.get(d, {}).get(room, {}).get(slot, [])
                label = get_entry_label(entries) if entries else ""
                cell = ws.cell(row, col, label)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                if wd_idx == 5:
                    cell.fill = FILL_SAT
                elif wd_idx == 6:
                    cell.fill = FILL_SUN
                cell.border = BORDER_ALL
                col += 1

        ws.row_dimensions[row].height = 45
        row += 1

    return row
